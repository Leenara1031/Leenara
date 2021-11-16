from openpyxl import Workbook
from openpyxl import load_workbook
import os.path

def writeExcel(dataList):

    dir = 'C:/lnr_scraping/download/'
    fileName = 'worleWeatherInfo.xlsx'

    isFile = os.path.isfile(dir + fileName)

    if isFile:
        wb = load_workbook(dir + fileName)
        ws = wb['new_sheet']
        ws[ws.max_row]
        ws.append(dataList)

        wb.save(dir + fileName)
        wb.close()

    else:
        wb = Workbook(dir + fileName)
        ws = wb.create_sheet('new_sheet')
        ws.append(dataList)

        wb.save(dir + fileName)
        wb.close()











